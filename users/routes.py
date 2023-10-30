import csv
import io
import json

from bson import ObjectId
import gridfs
import openpyxl
from flask import Blueprint, render_template, request, url_for, redirect, flash, session, Response

from flask_login import login_user, current_user, logout_user, login_required

import constants
from config import mongo
from users import utils
from users.forms import LoginForm, ManagerForm, EmployeeForm, PermissionForm, RoleForm
from users.user import User


login_bp = Blueprint(
    'login_bp', __name__,
    template_folder='templates'
)

export_bp = Blueprint(
    'export_bp', __name__,
    url_prefix="/export",
    template_folder='templates'
)

roles_bp = Blueprint(
    'roles_bp', __name__,
    url_prefix="/role",
    template_folder='templates'
)

permissions_bp = Blueprint(
    'permissions_bp', __name__,
    url_prefix="/permission",
    template_folder='users/templates'
)

admin_bp = Blueprint(
    'admin_bp', __name__,
    url_prefix="/admin",
    template_folder='users/templates'
)

employee_bp = Blueprint(
    'employee_bp', __name__,
    url_prefix="/employee",
    template_folder='users/templates'
)

manager_bp = Blueprint(
    'manager_bp', __name__,
    url_prefix="/manager",
    template_folder='users/templates'
)


@admin_bp.route('/dashboard', methods=('GET', 'POST'))
@utils.role_required(["admin"])
def admin_dashboard():
    employees = session.get('filtered_employees')
    if employees is None:
        employees = utils.get_employees()
    session.pop('filtered_employees', None)
    return render_template(
        'admin_dashboard.html',
        managers=utils.get_managers(),
        employees=employees,
        permissions=utils.get_permissions(),
        roles=utils.get_roles(),
        department_choices=constants.department_choices,
        designation_choices=constants.designation_choices)


@roles_bp.route('/create', methods=('GET', 'POST'))
@utils.role_required(["admin"])
def create_role():
    form = RoleForm()
    form.permissions.choices = [(permission.get('_id'), permission.get('name')) for permission in utils.get_permissions()]
    if request.method == 'POST':
        name = form.name.data
        existing_role = mongo.db.roles.find_one({'name': name})
        if existing_role:
            flash('Role name must be unique.', 'danger')
        else:
            permissions_ids = form.permissions.data
            mongo.db.roles.insert_one({'name': name, 'permissions_ids': [ObjectId(each) for each in permissions_ids]})
            return redirect(url_for('admin_bp.admin_dashboard'))
    return render_template('role_create.html', form=form)


@roles_bp.route('/delete/<role_id>', methods=('GET', 'POST'))
@utils.role_required(["admin"])
def delete_role(role_id):
    role_id = ObjectId(role_id)
    mongo.db.roles.delete_one({'_id': role_id})
    cursor_obj = mongo.db.users.find({"role_id": role_id}, {"_id": 1})
    [mongo.db.users.update_many({"manager": each["_id"]}, {"$set": {"manager": None}}) for each in cursor_obj]
    mongo.db.users.update_many(
        {"role_id": role_id},
        {"$set": {"role_id": None}}
    )
    return redirect(url_for('admin_bp.admin_dashboard'))


@permissions_bp.route('/create', methods=('GET', 'POST'))
@utils.role_required(["admin"])
def create_permission():
    form = PermissionForm()
    if form.validate_on_submit():
        name = request.form['name']
        existing_permission = mongo.db.permissions.find_one({'name': name})
        if existing_permission:
            flash('Permission name must be unique.', 'danger')
        else:
            mongo.db.permissions.insert_one({'name': name})
            return redirect(url_for('admin_bp.admin_dashboard'))
    return render_template('permission_create.html', form=form)


@permissions_bp.route('/delete/<permission_id>', methods=('GET', 'POST'))
@utils.role_required(["admin"])
def delete_permission(permission_id):
    mongo.db.roles.update_many({}, {"$pull": {"permissions_ids": permission_id}})
    mongo.db.permissions.delete_one({'_id': ObjectId(permission_id)})
    return redirect(url_for('admin_bp.admin_dashboard'))


@manager_bp.route('/dashboard', methods=('GET', 'POST'))   # Focus here
@utils.role_required(["manager"])
def manager_dashboard():
    employees = session.get('filtered_employees')
    if employees is None:
        employees = utils.get_employees()
    session.pop('filtered_employees', None)
    return render_template(
        'manager_dashboard.html',
        employees=employees, department_choices=constants.department_choices, designation_choices=constants.designation_choices)


@manager_bp.route('/create', methods=('GET', 'POST'))
@utils.role_required(["admin"])
def create_manager():
    if not utils.get_manager_role():
        flash("You need to first create a role with 'manager'", 'danger')
        if current_user.role == "manager":
            return redirect(url_for('manager_bp.manager_dashboard'))
        return redirect(url_for('admin_bp.admin_dashboard'))
    form = ManagerForm()
    if form.validate_on_submit():
        if utils.is_email_exists(request.form['email']):
            flash('Email must be unique.', 'danger')
        elif utils.is_phonenumber_exists(request.form['phone']):
            flash('Phone number must be unique.', 'danger')
        else:
            mongo.db.users.insert_one(
                {'name': request.form['name'], 'email': request.form['email'], 'phone': request.form['phone'],
                 'password': request.form['password'], 'hired_date': request.form['hired_date'],
                 'role_id': utils.get_manager_role()['_id']})
            if current_user.role == "manager":
                return redirect(url_for('manager_bp.manager_dashboard'))
            return redirect(url_for('admin_bp.admin_dashboard'))
    return render_template('manager_create.html', form=form)


@employee_bp.route('/dashboard', methods=('GET', 'POST'))
@utils.role_required(["employee"])
def employee_dashboard():
    user_id = current_user.id
    user_id = ObjectId(user_id)
    user = mongo.db.users.find_one({"_id": user_id})
    if user:
        user['_id'] = str(user['_id'])
        if user.get('manager'):
            user['manager'] = mongo.db.users.find_one({"_id": user["manager"]})["name"]
        else:
            user["manager"] = ""
        return render_template('employee_dashboard.html', user=user)
    flash('Employee details not found.', 'danger')
    return redirect(url_for('login_bp.login'))


@employee_bp.route('/detail/<employee_id>', methods=('GET', 'POST'))
@utils.role_required(["admin", "manager"])
def employee_detail(employee_id):
    form = EmployeeForm()
    form.manager.choices = [(manager.get('_id'), manager.get('name')) for manager in utils.get_managers()]
    if ObjectId.is_valid(employee_id):
        employee = mongo.db.users.find_one({'_id': ObjectId(employee_id)})
        if employee:
            return render_template('employee_detail.html', form=form, employee=employee)
    flash('Employee details not found.', 'danger')
    return redirect(url_for('login_bp.login'))


@employee_bp.route('/create', methods=('GET', 'POST'))
@utils.role_required(["admin", "manager"])
def create_employee():
    if not utils.get_employee_role():
        flash("You need to first create a role with 'employee'", 'danger')
        if current_user.role == "manager":
            return redirect(url_for('manager_bp.manager_dashboard'))
        return redirect(url_for('admin_bp.admin_dashboard'))
    # role_detail = utils.get_role_data(current_user.role)
    # if not 'admin_access' in role_detail.get('permission_names', []) and not 'add_employee' in role_detail.get('permisision_names', []):
    #     flash("You don't have permission to access this page.", 'danger')
    #     return redirect(url_for('login_bp.login'))
    form = EmployeeForm()
    form.manager.choices = [(manager.get('_id'), manager.get('name')) for manager in utils.get_managers()]
    if form.validate_on_submit():
        if utils.is_email_exists(request.form['email']):
            flash('Email must be unique.', 'danger')
        elif utils.is_phonenumber_exists(request.form['phone']):
            flash('Phone number must be unique.', 'danger')
        else:
            mongo.db.users.insert_one(
                {'name': request.form['name'], 'email': request.form['email'], 'phone': request.form['phone'],
                 'password': request.form['password'], 'designation': request.form['designation'], 'department': request.form['department'],
                 'manager': ObjectId(request.form['manager']), 'hired_date': request.form['hired_date'],
                 'role_id': utils.get_employee_role()['_id']})
            if current_user.role == "manager":
                return redirect(url_for('manager_bp.manager_dashboard'))
            return redirect(url_for('admin_bp.admin_dashboard'))
    return render_template('employee_create.html', form=form)


@employee_bp.route('/edit/<employee_id>', methods=('GET', 'POST'))
@utils.role_required(["admin", "manager"])
def edit_employee(employee_id):
    form = EmployeeForm()
    form.manager.choices = [(manager.get('_id'), manager.get('name')) for manager in utils.get_managers()]
    if request.method == 'POST':
        new_data = {
            'name': request.form['name'],
            'email': request.form['email'],
            'phone': request.form['phone'],
            'designation': request.form['designation'],
            'department': request.form['department'],
            'manager': ObjectId(request.form['manager']),
            'hired_date': request.form['hired_date']
        }
        mongo.db.users.update_one({'_id': ObjectId(employee_id)}, {'$set': new_data})
        if current_user.role == "manager":
            return redirect(url_for('manager_bp.manager_dashboard'))
        return redirect(url_for('admin_bp.admin_dashboard'))
    else:
        employee = mongo.db.users.find_one({'_id': ObjectId(employee_id)})
        return render_template('employee_edit.html', form=form, employee=employee)


@employee_bp.route('/delete/<employee_id>', methods=('GET', 'POST'))   # Focus here
@utils.role_required(["admin", "manager"])
def delete_employee(employee_id):
    mongo.db.users.delete_one({'_id': ObjectId(employee_id)})
    return redirect(url_for('manager_bp.manager_dashboard'))


@employee_bp.route('/filter_employees', methods=['POST'])
@utils.role_required(["admin", "manager"])
def filter_employees():
    designation = request.form.get('designation')
    department = request.form.get('department')
    employees = utils.get_employees()
    if designation and department:
        employees = [employee for employee in employees if
                     employee['designation'] == designation and employee['department'] == department]
    elif designation:
        employees = [employee for employee in employees if employee['designation'] == designation]
    elif department:
        employees = [employee for employee in employees if employee['department'] == department]
    for employee in employees:
        employee['_id'] = str(employee['_id'])
        employee['manager'] = str(employee['manager'])
        employee['role_id'] = str(employee['role_id'])
    session['filtered_employees'] = employees
    if current_user.role == "manager":
        return redirect(url_for('manager_bp.manager_dashboard'))
    return redirect(url_for('admin_bp.admin_dashboard'))


@employee_bp.route('/search_employees', methods=['GET', 'POST'])
@utils.role_required(["admin", "manager"])
def search_employees():
    search_query = request.form.get('search')
    employees = utils.get_employees()
    if search_query:
        employees = [employee for employee in employees if
                     (search_query in employee.get('name', '') or search_query in employee.get('phone', '') or
                      search_query in employee.get('email', ''))]
        if current_user.role == "manager":
            return render_template(
                'manager_dashboard.html',
                employees=employees,
                department_choices=constants.department_choices,
                designation_choices=constants.designation_choices)
        return render_template(
            'admin_dashboard.html',
            managers=utils.get_managers(),
            employees=employees,
            permissions=utils.get_permissions(),
            roles=utils.get_roles(),
            department_choices=constants.department_choices,
            designation_choices=constants.designation_choices)
    if current_user.role == "manager":
        return redirect(url_for('manager_bp.manager_dashboard'))
    return redirect(url_for('admin_bp.admin_dashboard'))


@export_bp.route('/<export_type>', methods=['GET', 'POST'])
@utils.role_required(["admin", "manager"])
def export(export_type):
    employees = utils.get_employees()
    fs = gridfs.GridFS(mongo.db)
    headers = ['Name', 'Email', 'Phone', 'Manager', 'Designation', 'Department']
    if export_type == 'csv':
        output = io.StringIO()
        csv_writer = csv.writer(output)
        csv_writer.writerow(headers)
        for employee in employees:
            csv_writer.writerow(
                [employee.get('name'), employee.get('email'), employee.get('phone'), employee.get('manager_name'),
                 employee.get('designation'), employee.get('department')])
        output.seek(0)
        csv_data = output.getvalue().encode('utf-8')
        fs.put(csv_data, filename='employees.csv')
        response = Response(
            output,
            content_type='text/csv',
            headers={'Content-Disposition': 'attachment; filename=employees.csv'}
        )
        return response
    elif export_type == 'json':
        employees = utils.modify_employee_data(employees)
        data_to_download = json.dumps(employees, indent=4)
        data_bytes = data_to_download.encode('utf-8')
        fs.put(data_bytes, filename='employees.json')
        response = Response(
            data_bytes,
            content_type='application/json',
            headers={'Content-Disposition': 'attachment; filename=employees.json'}
        )
        return response
    elif export_type == 'xlsx':
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        for col, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=col, value=header)
        for row, employee in enumerate(employees, 2):
            for col, header in enumerate(headers, 1):
                value = employee.get(header.lower()) if header != "Manager" else employee.get("manager_name", "")
                worksheet.cell(row=row, column=col, value=value)
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        xlsx_data = output.getvalue()
        fs.put(xlsx_data, filename="employees.xlsx")
        response = Response(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': 'attachment; filename=employees.xlsx' }
        )
        return response


@login_bp.route('/', methods=('GET', 'POST'))
def login():
    form = LoginForm()
    if form.validate_on_submit():
        username = form.username.data
        password = form.password.data
        user_data = mongo.db.users.find_one({"$or": [{"email": username}, {"phone": username}]})
        if user_data and user_data['password'] == password:
            user_role = mongo.db.roles.find_one({"_id": user_data.get('role_id')})
            if not user_data.get('role_id') or not user_role:
                flash('Login failed. No Role Assigned.', 'danger')
                return render_template('login.html', form=form)
            user_role = mongo.db.roles.find_one({"_id": user_data['role_id']})
            user = User(user_id=str(user_data["_id"]), username=user_data["name"], role=user_role["name"])
            login_user(user)
            if user_role.get('name') == "employee":
                return redirect(url_for('employee_bp.employee_dashboard'))
            elif user_role.get('name') == "manager":
                return redirect(url_for('manager_bp.manager_dashboard'))
            elif user_role.get('name') == "admin":
                return redirect(url_for('admin_bp.admin_dashboard'))
            else:
                flash('Invalid role!', 'danger')

            return redirect(url_for('admin_bp.admin_home'))
        else:
            flash('Login failed. Check your credentials.', 'danger')
    logout()
    return render_template('login.html', form=form)


@login_bp.route('/logout')
def logout():
    session.pop('username', None)
    session.pop('user_id', None)
    logout_user()
    return redirect(url_for('login_bp.login'))
